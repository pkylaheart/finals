import os
import tkinter as tk
from tkinter import ttk
from openpyxl import Workbook, load_workbook


def create_excel():
    if not os.path.exists("grades.xlsx"):
        wb = Workbook()
        ws = wb.active
        ws.title = "Grades Tracker"
        ws.append(["Name", "Course", "Score"])  
        wb.save("grades.xlsx")

def display_data():
    if not os.path.exists("grades.xlsx"):
        print("Excel file does not exist.")
        return

    wb = load_workbook("grades.xlsx")
    sheet = wb.active

    for item in treeview.get_children():
        treeview.delete(item)

    list_values = list(sheet.values)

    for col_index, col_name in enumerate(list_values[0]):
        treeview.heading(cols[col_index], text=col_name)

    for value_tuple in list_values[1:]:
        treeview.insert('', tk.END, values=value_tuple)

def submit():
    name = entry_name.get().strip()
    course = entry_course.get().strip()
    score = entry_score.get().strip()

    if not name or not course or not score:
        print("Fields Left Blank")
        return

    try:
        score = int(score)
    except ValueError:
        print("Invalid Score. Please enter a number.")
        return

    wb = load_workbook("grades.xlsx")
    ws = wb.active
    ws.append([name, course, score])
    wb.save("grades.xlsx")
    print("Added Successfully.")

    notif.config(text="Submitted", font=("Arial", 10), foreground="green")
    entry_name.delete(0, tk.END)
    entry_course.delete(0, tk.END)
    entry_score.delete(0, tk.END)
    entry_name.focus()
    app.after(2000, lambda: notif.config(text=""))
    display_data()


# GUI setup
app = tk.Tk()
app.title("Score Tracker")
app.resizable(False, False)
create_excel()

frame = tk.Frame(app)
frame.pack()

widgets_frame = ttk.LabelFrame(frame, text="Score Tracker")
widgets_frame.grid(row=0, column=0, padx=20, pady=10)

label_frame = tk.Frame(widgets_frame)
label_frame.grid(row=0, column=0)

entry_name = ttk.Entry(label_frame)
entry_name.insert(0, "Name")
entry_name.bind("<FocusIn>", lambda e: entry_name.delete("0", "end") if entry_name.get() == "Name" else None)
entry_name.bind("<FocusOut>", lambda e: entry_name.insert(0, "Name") if entry_name.get() == "" else None)
entry_name.grid(row=0, column=0, padx=5, pady=(0, 5), sticky="ew")

entry_course = ttk.Entry(label_frame)
entry_course.insert(0, "Course")
entry_course.bind("<FocusIn>", lambda e: entry_course.delete("0", "end") if entry_course.get() == "Course" else None)
entry_course.bind("<FocusOut>", lambda e: entry_course.insert(0, "Course") if entry_course.get() == "" else None)
entry_course.grid(row=1, column=0, padx=5, pady=(0, 5), sticky="ew")

entry_score = ttk.Entry(label_frame)
entry_score.insert(0, "Score")
entry_score.bind("<FocusIn>", lambda e: entry_score.delete("0", "end") if entry_score.get() == "Score" else None)
entry_score.bind("<FocusOut>", lambda e: entry_score.insert(0, "Score") if entry_score.get() == "" else None)
entry_score.grid(row=2, column=0, padx=5, pady=(0, 5), sticky="ew")

button = ttk.Button(label_frame, text="Submit", command=submit)
button.grid(row=3, column=0, sticky="nsew")

notif = ttk.Label(app, text="", font=("Arial", 12))
notif.pack(pady=10)

show_excel = ttk.Frame(frame)
show_excel.grid(row=0, column=1, pady=10)
show_scroll = ttk.Scrollbar(show_excel)
show_scroll.pack(side="right", fill="y")

cols = ("Name", "Course", "Score")
treeview = ttk.Treeview(show_excel, show="headings", yscrollcommand=show_scroll.set, columns=cols, height=13)
treeview.column("Name", width=100)
treeview.column("Course", width=100)
treeview.column("Score", width=50)
treeview.pack()
show_scroll.config(command=treeview.yview)

display_data()

app.mainloop()
